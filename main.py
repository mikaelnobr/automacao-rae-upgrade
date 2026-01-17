import streamlit as st
import os
import re
import json
import time
import tempfile
import gc
from io import BytesIO

# --- 1. CONFIGURAÇÃO DE PÁGINA ---
st.set_page_config(
    page_title="Automação RAE CAIXA",
    page_icon="🏛️",
    layout="centered"
)

# --- 2. BANCO DE DADOS DE PROFISSIONAIS ---
PROFISSIONAIS = {
    "FRANCISCO DAVID MENESES DOS SANTOS": {
        "empresa": "FRANCISCO DAVID MENESES DOS SANTOS - F. D. MENESES DOS SANTOS",
        "cnpj": "54.801.096/0001-16",
        "cpf_emp": "058.756.003-73",
        "nome_resp": "FRANCISCO DAVID MENESES DOS SANTOS",
        "cpf_resp": "058.756.003-73",
        "registro": "336241CE"
    },
    "PALLOMA TEIXEIRA DA SILVA": {
        "empresa": "PALLOMA TEIXEIRA DA SILVA - PALLOMA TEIXEIRA ARQUITETURA LTDA",
        "cnpj": "54.862.474/0001-71",
        "cpf_emp": "064.943.593-10",
        "nome_resp": "PALLOMA TEIXEIRA DA SILVA",
        "cpf_resp": "064.943.593-10",
        "registro": "A184355-9"
    },
    "SANDY PEREIRA CORDEIRO": {
        "empresa": "SANDY PEREIRA CORDEIRO - CS ENGENHARIA",
        "cnpj": "54.794.898/0001-46",
        "cpf_emp": "071.222.553-60",
        "nome_resp": "SANDY PEREIRA CORDEIRO",
        "cpf_resp": "071.222.553-60",
        "registro": "356882CE"
    },
    "TIAGO VICTOR DE SOUSA": {
        "empresa": "TIAGO VICTOR DE SOUSA - T V S ENGENHARIA E ASSESSORIA",
        "cnpj": "54.806.521/0001-60",
        "cpf_emp": "068.594.803-00",
        "nome_resp": "TIAGO VICTOR DE SOUSA",
        "cpf_resp": "068.594.803-00",
        "registro": "346856CE"
    }
}

# --- 3. FUNÇÕES DE SUPORTE ---
def to_f(v):
    try:
        if v is None or v == "":
            return 0
        v = str(v).replace("R$", "").replace("%", "").replace(" ", "")
        if "," in v and "." in v:
            v = v.replace(".", "").replace(",", ".")
        elif "," in v:
            v = v.replace(",", ".")
        v = re.sub(r"[^\d.]", "", v)
        return float(v)
    except:
        return 0


def limpar_texto_para_ia(texto, limite=12000):
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = texto.strip()

    if len(texto) > limite:
        metade = limite // 2
        texto = texto[:metade] + "\n...\n" + texto[-metade:]

    return texto


def call_gemini(api_key, prompt):
    import google.generativeai as genai
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    for _ in range(3):
        try:
            resp = model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    response_mime_type="application/json",
                    temperature=0.1
                )
            )
            return json.loads(resp.text)
        except:
            time.sleep(2)

    return None


# --- 4. EXTRAÇÃO COM DOCLING (OTIMIZADA E SEGURA) ---
def extrair_com_docling(doc, nome_doc):
    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.datamodel.base_models import InputFormat
    from docling.backend.pypdfium2_backend import PyPdfiumDocumentBackend

    st.write(f"📂 Processando {nome_doc}...")

    pipeline = PdfPipelineOptions()
    pipeline.do_ocr = (nome_doc != "LAUDO")
    pipeline.do_table_structure = (nome_doc == "PLS")

    converter = DocumentConverter(
        allowed_formats=[InputFormat.PDF],
        format_options={
            InputFormat.PDF: PdfFormatOption(
                pipeline_options=pipeline,
                backend=PyPdfiumDocumentBackend
            )
        }
    )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(doc.getbuffer())
        path = tmp.name

    try:
        res = converter.convert(path)
        texto = res.document.export_to_markdown()

        del res
        del converter
        gc.collect()

        return limpar_texto_para_ia(texto)

    finally:
        if os.path.exists(path):
            os.remove(path)


# --- 5. APP PRINCIPAL ---
def main():
    st.title("🏛️ Automação RAE CAIXA")
    st.markdown("##### v6.1 — Estável para Laudo + PLS + Alvará")

    with st.sidebar:
        st.header("⚙️ Configurações")
        api_key = st.text_input("Gemini API Key:", type="password")
        st.divider()
        st.subheader("👤 Responsável Técnico")
        resp = st.selectbox("Selecione o Profissional:", list(PROFISSIONAIS.keys()))
        st.divider()
        st.caption("Backend PyPdfium2 | Processamento sequencial")

    st.subheader("📂 Documentação")
    col1, col2 = st.columns(2)
    with col1:
        pdf_laudo = st.file_uploader("1. Laudo Técnico (PDF)", type=["pdf"])
        pdf_pls = st.file_uploader("3. PLS (PDF)", type=["pdf"])
    with col2:
        excel_template = st.file_uploader("2. Modelo RAE (.xlsm)", type=["xlsm"])
        pdf_alvara = st.file_uploader("4. Alvará (PDF)", type=["pdf"])

    if st.button("🚀 INICIAR PROCESSAMENTO"):
        if not api_key or not pdf_laudo or not excel_template:
            st.warning("Preencha os campos obrigatórios.")
            return

        try:
            with st.status("Extraindo documentos...", expanded=True):
                textos = []

                textos.append(f"\n--- LAUDO ---\n{extrair_com_docling(pdf_laudo, 'LAUDO')}")
                gc.collect()

                if pdf_pls:
                    textos.append(f"\n--- PLS ---\n{extrair_com_docling(pdf_pls, 'PLS')}")
                    gc.collect()

                if pdf_alvara:
                    textos.append(f"\n--- ALVARÁ ---\n{extrair_com_docling(pdf_alvara, 'ALVARA')}")
                    gc.collect()

                texto_total = "\n".join(textos)
                del textos
                gc.collect()

            st.write("🧠 IA: Cruzando informações...")
            prompt = f"""
            Você é um engenheiro revisor da CAIXA.
            Gere um JSON estrito.

            CAMPOS:
            valor_imovel
            contratacao
            percentual_pls
            acumulado_pls
            lat_s
            long_w

            TEXTO:
            {texto_total}
            """

            dados = call_gemini(api_key, prompt)
            if not dados:
                st.error("Falha na resposta da IA.")
                return

            st.write("📊 Preenchendo Excel...")
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(excel_template.read()), keep_vba=True)

            if "RAE" in wb.sheetnames:
                ws = wb["RAE"]
                ws["AH63"] = dados.get("contratacao", "")
                ws["AH66"] = to_f(dados.get("valor_imovel", 0))
                ws["W93"] = to_f(dados.get("percentual_pls", 0))

                prof = PROFISSIONAIS[resp]
                ws["I315"] = prof["empresa"].upper()
                ws["I316"] = prof["cnpj"]
                ws["AE315"] = prof["nome_resp"].upper()
                ws["AO316"] = prof["registro"]

            output = BytesIO()
            wb.save(output)

            st.success("✅ RAE gerada com sucesso!")
            st.download_button(
                "📥 BAIXAR RAE FINAL",
                data=output.getvalue(),
                file_name="RAE_FINAL.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
            )

        except Exception as e:
            st.error(f"Erro inesperado: {e}")
            gc.collect()


if __name__ == "__main__":
    main()
